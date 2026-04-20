# -*- coding: utf-8 -*-
"""
Parameter Transformer MEP (v2)
- Export į CSV/XLSX (Excel roundtrip)
- Import iš CSV/XLSX su preview
- Rollback log (JSON) + atstatymas iš log
"""

from pyrevit import revit, DB, forms
import os
import csv
import json
import codecs
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


doc = revit.doc


CATEGORY_MAP = {
    u"Pipes": DB.BuiltInCategory.OST_PipeCurves,
    u"Ducts": DB.BuiltInCategory.OST_DuctCurves,
    u"Pipe Fittings": DB.BuiltInCategory.OST_PipeFitting,
    u"Duct Fittings": DB.BuiltInCategory.OST_DuctFitting,
    u"Pipe Accessories": DB.BuiltInCategory.OST_PipeAccessory,
    u"Duct Accessories": DB.BuiltInCategory.OST_DuctAccessory,
    u"Cable Trays": DB.BuiltInCategory.OST_CableTray,
    u"Mechanical Equipment": DB.BuiltInCategory.OST_MechanicalEquipment,
    u"Plumbing Fixtures": DB.BuiltInCategory.OST_PlumbingFixtures,
    u"Duct Terminals": DB.BuiltInCategory.OST_DuctTerminal,
    u"Conduits": DB.BuiltInCategory.OST_Conduit,
    u"Conduit Fittings": DB.BuiltInCategory.OST_ConduitFitting,
}

BASE_COLS = [u"UniqueId", u"ElementId", u"Category", u"Family", u"Type", u"Level", u"System"]


def safe_text(val):
    if val is None:
        return u""
    try:
        return unicode(val)
    except Exception:
        try:
            return str(val).decode('utf-8', 'ignore')
        except Exception:
            try:
                return str(val)
            except Exception:
                return u""


def default_path(filename):
    home = os.path.expanduser("~")
    return os.path.join(home, filename)


def get_ext(path):
    return os.path.splitext(path)[1].lower()


def get_level_name(elem):
    try:
        p = elem.get_Parameter(DB.BuiltInParameter.FAMILY_LEVEL_PARAM)
        if p and p.HasValue:
            return safe_text(p.AsValueString() or p.AsString())
    except Exception:
        pass

    try:
        p = elem.get_Parameter(DB.BuiltInParameter.RBS_START_LEVEL_PARAM)
        if p and p.HasValue:
            return safe_text(p.AsValueString() or p.AsString())
    except Exception:
        pass

    return u""


def get_system_name(elem):
    try:
        p = elem.get_Parameter(DB.BuiltInParameter.RBS_SYSTEM_NAME_PARAM)
        if p and p.HasValue:
            return safe_text(p.AsString() or p.AsValueString())
    except Exception:
        pass
    return u""


def get_family_type(elem):
    fam = u""
    typ = u""
    try:
        et = doc.GetElement(elem.GetTypeId())
        if et:
            if hasattr(et, 'FamilyName') and et.FamilyName:
                fam = safe_text(et.FamilyName)
            if hasattr(et, 'Name') and et.Name:
                typ = safe_text(et.Name)
    except Exception:
        pass
    return fam, typ


def collect_elements(selected_category_names, system_filter_text, level_filter_text):
    bics = [CATEGORY_MAP[n] for n in selected_category_names]
    filter_cats = DB.ElementMulticategoryFilter(bics)
    elements = DB.FilteredElementCollector(doc).WherePasses(filter_cats).WhereElementIsNotElementType().ToElements()

    out = []
    for e in elements:
        sys_name = get_system_name(e)
        lvl_name = get_level_name(e)

        if system_filter_text and system_filter_text.lower() not in sys_name.lower():
            continue
        if level_filter_text and level_filter_text.lower() not in lvl_name.lower():
            continue

        out.append(e)

    return out


def get_editable_param_names(elements):
    names = set()
    for e in elements:
        try:
            for p in e.Parameters:
                try:
                    if p and p.Definition and p.Definition.Name and (not p.IsReadOnly):
                        names.add(safe_text(p.Definition.Name))
                except Exception:
                    pass
        except Exception:
            pass

    names = sorted(list(names))
    if len(names) > 200:
        names = names[:200]
    return names


def get_param_display_value(elem, param_name):
    p = elem.LookupParameter(param_name)
    if not p:
        return u""

    try:
        st = p.StorageType
        if st == DB.StorageType.String:
            return safe_text(p.AsString() or u"")
        if st == DB.StorageType.Double:
            return safe_text(p.AsValueString() or u"")
        if st == DB.StorageType.Integer:
            return safe_text(p.AsInteger())
        if st == DB.StorageType.ElementId:
            eid = p.AsElementId()
            if eid:
                return safe_text(eid.IntegerValue)
            return u""
    except Exception:
        pass

    try:
        return safe_text(p.AsValueString() or p.AsString() or u"")
    except Exception:
        return u""


def set_param_from_text(elem, param_name, new_text):
    p = elem.LookupParameter(param_name)
    if (not p) or p.IsReadOnly:
        return False, u"Parametras nerastas arba read-only"

    txt = safe_text(new_text).strip()

    try:
        st = p.StorageType

        if st == DB.StorageType.String:
            p.Set(txt)
            return True, u""

        if st == DB.StorageType.Double:
            try:
                ok = p.SetValueString(txt)
                if ok:
                    return True, u""
            except Exception:
                pass
            try:
                val = float(txt.replace(',', '.'))
                p.Set(val)
                return True, u""
            except Exception:
                return False, u"Neteisinga skaitinė reikšmė"

        if st == DB.StorageType.Integer:
            try:
                p.Set(int(float(txt.replace(',', '.'))))
                return True, u""
            except Exception:
                return False, u"Neteisinga sveiko skaičiaus reikšmė"

        if st == DB.StorageType.ElementId:
            try:
                p.Set(DB.ElementId(int(float(txt.replace(',', '.')))))
                return True, u""
            except Exception:
                return False, u"Neteisingas ElementId"

        return False, u"Nepalaikomas parametro tipas"

    except Exception as ex:
        return False, safe_text(ex)


def export_table_rows(elements, selected_params):
    headers = BASE_COLS + selected_params
    rows = []

    for e in elements:
        fam, typ = get_family_type(e)
        row = {
            u"UniqueId": safe_text(e.UniqueId),
            u"ElementId": safe_text(e.Id.IntegerValue),
            u"Category": safe_text(e.Category.Name if e.Category else u""),
            u"Family": fam,
            u"Type": typ,
            u"Level": get_level_name(e),
            u"System": get_system_name(e)
        }
        for pn in selected_params:
            row[pn] = get_param_display_value(e, pn)

        rows.append(row)

    return headers, rows


def write_csv(path, headers, rows):
    with codecs.open(path, 'w', 'utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([safe_text(r.get(h, u"")) for h in headers])


def read_csv(path):
    with codecs.open(path, 'r', 'utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = [dict((safe_text(k), safe_text(v)) for k, v in rr.items()) for rr in reader]
        headers = [safe_text(h) for h in (reader.fieldnames or [])]
    return headers, rows


def write_xlsx(path, headers, rows):
    if not HAS_OPENPYXL:
        raise Exception(u"openpyxl nėra prieinamas šioje aplinkoje")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MEP_Parameters"

    ws.append([safe_text(h) for h in headers])
    for r in rows:
        ws.append([safe_text(r.get(h, u"")) for h in headers])

    wb.save(path)


def read_xlsx(path):
    if not HAS_OPENPYXL:
        raise Exception(u"openpyxl nėra prieinamas šioje aplinkoje")

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    headers = [safe_text(h) for h in all_rows[0] if h is not None]
    rows = []

    for rr in all_rows[1:]:
        if rr is None:
            continue
        data = {}
        empty_row = True
        for i, h in enumerate(headers):
            val = rr[i] if i < len(rr) else u""
            txt = safe_text(val)
            if txt.strip():
                empty_row = False
            data[h] = txt
        if not empty_row:
            rows.append(data)

    return headers, rows


def write_table(path, headers, rows):
    ext = get_ext(path)
    if ext == '.xlsx':
        write_xlsx(path, headers, rows)
    else:
        write_csv(path, headers, rows)


def read_table(path):
    ext = get_ext(path)
    if ext == '.xlsx':
        return read_xlsx(path)
    return read_csv(path)


def create_rollback_log(import_path, applied_changes):
    if not applied_changes:
        return None

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    folder = os.path.dirname(import_path) or os.path.expanduser('~')
    log_path = os.path.join(folder, 'parameter_transformer_rollback_{}.json'.format(ts))

    payload = {
        "version": "v2",
        "created_at": datetime.now().isoformat(),
        "source_file": import_path,
        "changes": applied_changes
    }

    with codecs.open(log_path, 'w', 'utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    return log_path


def import_table(path):
    headers, rows = read_table(path)

    editable_cols = [h for h in headers if h not in BASE_COLS]

    if u"UniqueId" not in headers:
        forms.alert(u"Faile trūksta 'UniqueId' stulpelio.")
        return

    changes_preview = []
    change_plan = []

    for r in rows:
        uid = safe_text(r.get(u"UniqueId", u"")).strip()
        if not uid:
            continue

        elem = doc.GetElement(uid)
        if not elem:
            continue

        for pn in editable_cols:
            if pn is None:
                continue
            new_val = safe_text(r.get(pn, u""))
            old_val = get_param_display_value(elem, pn)
            if new_val == old_val:
                continue

            change_plan.append((elem, pn, new_val, old_val, uid))
            if len(changes_preview) < 30:
                changes_preview.append(u"{} | {}: '{}' -> '{}'".format(elem.Id.IntegerValue, pn, old_val, new_val))

    if not change_plan:
        forms.alert(u"Faile nerasta jokių pakeitimų pritaikymui.")
        return

    msg = u"Rasti pakeitimai: {}\n\nPeržiūra (iki 30):\n{}\n\nTaikyti pakeitimus?".format(
        len(change_plan),
        u"\n".join(changes_preview)
    )

    decision = forms.CommandSwitchWindow.show([u"Taikyti", u"Atšaukti"], message=msg)
    if decision != u"Taikyti":
        forms.alert(u"Importas atšauktas.")
        return

    ok_count = 0
    fail_count = 0
    fail_msgs = []
    applied_changes = []

    with revit.Transaction(u"Parameter Transformer MEP - Import"):
        for elem, pn, new_val, old_val, uid in change_plan:
            ok, err = set_param_from_text(elem, pn, new_val)
            if ok:
                ok_count += 1
                applied_changes.append({
                    "unique_id": safe_text(uid),
                    "element_id": safe_text(elem.Id.IntegerValue),
                    "param": safe_text(pn),
                    "old_value": safe_text(old_val),
                    "new_value": safe_text(new_val)
                })
            else:
                fail_count += 1
                if len(fail_msgs) < 20:
                    fail_msgs.append(u"{} | {} -> {}".format(elem.Id.IntegerValue, pn, err))

    log_path = create_rollback_log(path, applied_changes)

    result = [u"Pritaikyta: {}".format(ok_count), u"Nepavyko: {}".format(fail_count)]
    if log_path:
        result.append(u"Rollback log: {}".format(log_path))
    if fail_msgs:
        result.append(u"\nKlaidų pavyzdžiai:\n" + u"\n".join(fail_msgs))

    forms.alert(u"\n".join(result))


def rollback_from_log(log_path):
    if not os.path.exists(log_path):
        forms.alert(u"Rollback log failas nerastas:\n{}".format(log_path))
        return

    with codecs.open(log_path, 'r', 'utf-8') as f:
        payload = json.load(f)

    changes = payload.get('changes', [])
    if not changes:
        forms.alert(u"Rollback log tuščias.")
        return

    preview = []
    for c in changes[:30]:
        preview.append(u"{} | {}: '{}' <- '{}'".format(
            safe_text(c.get('element_id', u'?')),
            safe_text(c.get('param', u'?')),
            safe_text(c.get('old_value', u'')),
            safe_text(c.get('new_value', u''))
        ))

    msg = u"Rollback įrašų: {}\n\nPeržiūra (iki 30):\n{}\n\nAtstatyti?".format(len(changes), u"\n".join(preview))
    decision = forms.CommandSwitchWindow.show([u"Atstatyti", u"Atšaukti"], message=msg)
    if decision != u"Atstatyti":
        return

    ok_count = 0
    fail_count = 0
    fail_msgs = []

    with revit.Transaction(u"Parameter Transformer MEP - Rollback"):
        for c in changes:
            uid = safe_text(c.get('unique_id', u'')).strip()
            pn = safe_text(c.get('param', u'')).strip()
            old_val = safe_text(c.get('old_value', u''))

            if not uid or not pn:
                fail_count += 1
                continue

            elem = doc.GetElement(uid)
            if not elem:
                fail_count += 1
                if len(fail_msgs) < 20:
                    fail_msgs.append(u"Elementas nerastas: {}".format(uid))
                continue

            ok, err = set_param_from_text(elem, pn, old_val)
            if ok:
                ok_count += 1
            else:
                fail_count += 1
                if len(fail_msgs) < 20:
                    fail_msgs.append(u"{} | {} -> {}".format(elem.Id.IntegerValue, pn, err))

    result = [u"Rollback atstatyta: {}".format(ok_count), u"Rollback nepavyko: {}".format(fail_count)]
    if fail_msgs:
        result.append(u"\nKlaidų pavyzdžiai:\n" + u"\n".join(fail_msgs))
    forms.alert(u"\n".join(result))


def export_flow():
    selected_cats = forms.SelectFromList.show(
        sorted(CATEGORY_MAP.keys()),
        title=u"Pasirink kategorijas eksportui",
        multiselect=True
    )
    if not selected_cats:
        return

    system_filter = forms.ask_for_string(
        default=u"",
        prompt=u"Sistemos filtras (contains). Tuščia = visi:",
        title=u"Filtras pagal sistemą"
    ) or u""

    level_filter = forms.ask_for_string(
        default=u"",
        prompt=u"Level filtras (contains). Tuščia = visi:",
        title=u"Filtras pagal level"
    ) or u""

    elements = collect_elements(selected_cats, system_filter, level_filter)
    if not elements:
        forms.alert(u"Pagal pasirinktus filtrus elementų nerasta.")
        return

    param_names = get_editable_param_names(elements)
    selected_params = forms.SelectFromList.show(
        param_names,
        title=u"Pasirink parametrus eksportui (Excel redagavimui)",
        multiselect=True
    )
    if not selected_params:
        forms.alert(u"Nepasirinkta parametrų eksportui.")
        return

    fmt = forms.CommandSwitchWindow.show([u"CSV", u"XLSX"], message=u"Pasirink eksporto formatą")
    if not fmt:
        return

    if fmt == u"XLSX" and (not HAS_OPENPYXL):
        forms.alert(u"XLSX eksportui reikia openpyxl bibliotekos. Naudok CSV arba įdiek openpyxl.")
        return

    ext = '.xlsx' if fmt == u"XLSX" else '.csv'
    default_out = default_path('mep_parameter_transformer_export{}'.format(ext))

    out_path = forms.ask_for_string(
        default=default_out,
        prompt=u"Eksporto failo kelias ({}):".format(fmt),
        title=u"Export {}".format(fmt)
    )
    if not out_path:
        return

    headers, rows = export_table_rows(elements, selected_params)

    try:
        write_table(out_path, headers, rows)
        forms.alert(u"Eksportuota elementų: {}\nFailas: {}".format(len(rows), out_path))
    except Exception as ex:
        forms.alert(u"Eksportas nepavyko:\n{}".format(safe_text(ex)))


def import_flow():
    default_in = default_path('mep_parameter_transformer_export.csv')
    in_path = forms.ask_for_string(
        default=default_in,
        prompt=u"Importo failo kelias (.csv arba .xlsx):",
        title=u"Import"
    )
    if not in_path:
        return

    if not os.path.exists(in_path):
        forms.alert(u"Failas nerastas:\n{}".format(in_path))
        return

    if get_ext(in_path) == '.xlsx' and (not HAS_OPENPYXL):
        forms.alert(u"XLSX importui reikia openpyxl bibliotekos. Naudok CSV arba įdiek openpyxl.")
        return

    try:
        import_table(in_path)
    except Exception as ex:
        forms.alert(u"Importas nepavyko:\n{}".format(safe_text(ex)))


def rollback_flow():
    default_log = default_path('parameter_transformer_rollback_YYYYMMDD_HHMMSS.json')
    log_path = forms.ask_for_string(
        default=default_log,
        prompt=u"Rollback log (.json) kelias:",
        title=u"Rollback"
    )
    if not log_path:
        return

    try:
        rollback_from_log(log_path)
    except Exception as ex:
        forms.alert(u"Rollback nepavyko:\n{}".format(safe_text(ex)))


def main():
    action = forms.CommandSwitchWindow.show(
        [u"Export -> CSV/XLSX", u"Import <- CSV/XLSX", u"Rollback <- LOG"],
        message=u"Parameter Transformer MEP v2"
    )

    if action == u"Export -> CSV/XLSX":
        export_flow()
    elif action == u"Import <- CSV/XLSX":
        import_flow()
    elif action == u"Rollback <- LOG":
        rollback_flow()


if __name__ == '__main__':
    main()
